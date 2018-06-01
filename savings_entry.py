import openpyxl as op
from openpyxl.utils import coordinate_from_string, column_index_from_string
from search import find_header, find_name, find_net
wb = op.load_workbook('savings.xlsx')
wb_1 = op.load_workbook('DATA13.xlsx')


def show_current_savings(sheet_name, emp_name, wb):
    ''' show in console current data '''
    sheet = wb.get_sheet_by_name(sheet_name)
    header = find_header(sheet_name, wb)
    name_row = find_name(sheet_name, emp_name, wb)
    savings_dict ={}

    for temp in list(header.keys()):
        print(temp + ' : ' + str(sheet.cell(row = name_row, 
            column = header[temp]).value) or 0 + '\n')
        savings_dict[temp] = sheet.cell(row = name_row, 
            column = header[temp]).value
    return savings_dict

def data_entry(sheet_name, emp_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    header = find_header(sheet_name, wb)
    name_row = find_name(sheet_name, emp_name, wb)

    for temp in list(header.keys()):
        if temp not in ['NAME'] :
            
            try:
                str_in = input( '\n' + temp + ' :')
            except SyntaxError:
                str_in = None

            print (str_in)

            if (str_in is None) or (str_in.strip() == ''):
                print(sheet.cell(row = name_row, 
                    column = header[temp]).value)
            else:
                print('\n\n')
                try:
                    if temp == 'PAN':
                        sheet.cell(row = name_row,
                            column = header['PAN']).value = str_in
                    else : 
                        sheet.cell(row = name_row,
                            column = header[temp]).value = int(str_in)
                except ValueError:
                    print('Enter valid input')
                    break



    wb.save(filename = 'savings.xlsx')

def yearwise_calculation(sheet_name):
    sheet = wb.get_sheet_by_name('Sheet2')
    emp_name = list(sheet.columns)[0]
    header = list(sheet.rows)[0]
    hd_dic = {}

    #print(header)
    for counter in header:
        hd_dic[counter.value.strip()] =column_index_from_string(counter.column)





    for counter in emp_name:
        dic_data = {}
        total = {}
        row_val = 0

        if counter.value != 'NAME':

            
            print(counter.value)
            #for title in header:
            

                         
            for rw in sheet.iter_rows():
                for cl in rw:
                    if cl.value == counter.value:
                       row_val = cl.row

            for sh in wb_1.sheetnames:

                    
                try:
                    row_num1 = find_name(sh, counter.value, wb_1)
                    dic_data = find_net(sh, row_num1, wb_1)

                        
                except ValueError:
                    print("Name not found")
                        
                        

                for val in list(dic_data.keys()):
                    
                            
                    try:
                        total[val] +=dic_data[val]
                            
                    except KeyError:
                        total = dic_data
                        break

                    except TypeError:
                        pass

                dic_data = {}

                        
            # print(total)
            # print(hd_dic)

            for title in hd_dic.keys():
                if title not in ['NAME','PAN']:
                    sheet.cell(row = row_val, 
                    column = hd_dic[title]).value = total[title]


    wb.save(filename = 'savings.xlsx')
            


def main2():
    str_in = input('Enter Name : ')
    show_current_savings('Sheet1',str_in, wb)
    data_entry('Sheet1',str_in)




            


        
        

if __name__ == "__main__":
    main2()

